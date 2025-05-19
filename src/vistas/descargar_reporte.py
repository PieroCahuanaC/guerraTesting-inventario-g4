import tkinter as tk
import os 
from tkinter import messagebox

import pandas as pd  # Para generar reporte Excel
from reportlab.lib.pagesizes import letter  # Para definir tamaño del PDF
from reportlab.pdfgen import canvas  # Generador de PDFs (no usado directamente en esta versión)

from services.supabase_service import supabase  # Cliente de conexión a Supabase
from datetime import datetime  # Para obtener fecha y hora actual

#para mostrar mensaje de error personalizado si no hay internet
def mostrar_error_conexion_si_aplica(e, contexto="realizar la acción"):
    mensaje = str(e).lower()
    if any(palabra in mensaje for palabra in [
        "getaddrinfo failed",
        "failed to establish",
        "name or service not known",
        "temporary failure in name resolution"
    ]):
        messagebox.showerror("Error de conexión",
                             f"No se pudo {contexto} porque no hay conexión a internet.\nVerifica tu red e inténtalo nuevamente.")
    else:
        messagebox.showerror("Error inesperado", f"No se pudo {contexto}:\n{e}")


def crear_frame_reporte(root):
    """
    Crea un frame de interfaz gráfica para generar reportes en Excel y PDF
    con la información del inventario de productos almacenada en Supabase.
    """
    frame = tk.Frame(root, bg="white")

    # Título del frame
    titulo = tk.Label(frame, text="Descargar Reporte de Inventario", font=("Sans-serif", 32, "bold"), bg="white")
    titulo.pack(pady=60)

    # --- FUNCIÓN: Generar reporte Excel ---
    def generar_excel():
        """
        Consulta la base de datos y genera un archivo Excel (.xlsx) 
        con los datos del inventario.
        """
        try:
            # Consultar datos con relaciones (joins) a unidades y categorías
            response = supabase.table("productos").select(
                "id_producto, nombre, cantidad, precio, unidades(nombre_unidad), categorias(nombre_categoria)"
            ).execute()

            productos = response.data
            if not productos:
                messagebox.showinfo("Sin datos", "No hay productos en el inventario para generar el reporte Excel.")
                return

            productos = sorted(productos, key=lambda p: p["id_producto"])  # Ordenar por ID

            # Procesar datos con ID visual secuencial
            datos_limpios = []
            for idx, p in enumerate(productos, start=1):
                datos_limpios.append({
                    "N°.": idx,
                    "ID": p["id_producto"],
                    "Nombre": p["nombre"],
                    "Cantidad": p["cantidad"],
                    "Precio": p["precio"],
                    "Unidad": p["unidades"]["nombre_unidad"],
                    "Categoría": p["categorias"]["nombre_categoria"]
                })

            # Crear DataFrame y guardar como archivo Excel
            df = pd.DataFrame(datos_limpios)
            ruta_salida = os.path.abspath(os.path.join(os.path.dirname(__file__), "../../reporte_inventario.xlsx"))
            df.to_excel(ruta_salida, index=False)

            # --- APLICAR ESTILOS CON openpyxl ---
            from openpyxl import load_workbook
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

            wb = load_workbook(ruta_salida)
            ws = wb.active

            # Insertar fecha como primera fila
            ws.insert_rows(1)
            fecha_actual = datetime.now().strftime("%d/%m/%Y a las %H:%M")
            ws["A1"] = f"Reporte generado el {fecha_actual}"
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

            # Estilos
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill("solid", fgColor="4CAF50")
            id_fill = PatternFill("solid", fgColor="B3E5FC")  # celeste claro

            border = Border(left=Side(style='thin'), right=Side(style='thin'),
                            top=Side(style='thin'), bottom=Side(style='thin'))
            center_align = Alignment(horizontal="center", vertical="center")

            # Aplicar estilos a encabezado
            # Aplicar estilos a encabezado con color especial para columna "N°."
            for cell in ws[2]:
                cell.font = header_font
                cell.border = border
                cell.alignment = center_align
                if cell.value == "N°.":
                    cell.fill = PatternFill("solid", fgColor="1976D2")  # Azul oscuro
                else:
                    cell.fill = header_fill  # Verde original


            # Aplicar estilos a datos
            for row in ws.iter_rows(min_row=3, max_row=ws.max_row, max_col=ws.max_column):
                for cell in row:
                    cell.border = border
                    cell.alignment = center_align

            # Ajuste de anchos
            ws.column_dimensions["A"].width = 10  # N°.
            ws.column_dimensions["B"].width = 10  # ID
            ws.column_dimensions["C"].width = 35  # Nombre
            ws.column_dimensions["D"].width = 26  # Cantidad
            ws.column_dimensions["E"].width = 26  # Precio
            ws.column_dimensions["F"].width = 26  # Unidad
            ws.column_dimensions["G"].width = 25  # Categoría

            wb.save(ruta_salida)

            messagebox.showinfo("Éxito", "Reporte Excel generado correctamente en la carpeta del proyecto.")
        except Exception as e:
            mostrar_error_conexion_si_aplica(e, "generar el reporte Excel")

    # Botón para generar Excel
    btn_excel = tk.Button(frame, text="Generar y Guardar Excel", command=generar_excel,
                          font=("Sans-serif", 18, "bold"), bg="#4CAF50", fg="white", height=2)
    btn_excel.pack(pady=30)

    # --- FUNCIÓN: Generar reporte PDF ---
    def generar_pdf():
        """
        Consulta la base de datos y genera un archivo PDF con los datos
        del inventario organizados en una tabla.
        """
        try:
            response = supabase.table("productos").select(
                "id_producto, nombre, cantidad, precio, unidades(nombre_unidad), categorias(nombre_categoria)"
            ).execute()

            productos = response.data

            if not productos:
                messagebox.showinfo("Sin datos", "No hay productos en el inventario para generar el reporte PDF.")
                return

            productos = sorted(productos, key=lambda p: p["id_producto"])

            datos_limpios = [
                ["N°.", "ID", "Nombre", "Cantidad", "Precio (S/.)", "Unidad", "Categoría"]
            ]
            for idx, p in enumerate(productos, start=1):
                datos_limpios.append([
                    idx,
                    p["id_producto"],
                    p["nombre"],
                    p["cantidad"],
                    f"{p['precio']:.2f}",
                    p["unidades"]["nombre_unidad"],
                    p["categorias"]["nombre_categoria"]
                ])


            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet

            ruta_pdf = os.path.abspath(os.path.join(os.path.dirname(__file__), "../../reporte_inventario.pdf"))
            doc = SimpleDocTemplate(ruta_pdf, pagesize=letter)

            elementos = []
            styles = getSampleStyleSheet()

            titulo = Paragraph("Reporte de Inventario - Muebles Moderno", styles["Title"])
            elementos.append(titulo)
            elementos.append(Spacer(1, 20))

            fecha_actual = datetime.now().strftime("%d/%m/%Y a las %H:%M")
            fecha_parrafo = Paragraph(f"Reporte generado el {fecha_actual}", styles["Normal"])
            elementos.append(fecha_parrafo)
            elementos.append(Spacer(1, 20))

            tabla = Table(datos_limpios, repeatRows=1)
            tabla.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.gray),  # Fila de encabezado
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 12),

                ("BACKGROUND", (0, 1), (-1, -1), colors.beige),  # Cuerpo general

                ("BACKGROUND", (0, 0), (0, -1), colors.steelblue),  # Columna "N°." completa
                ("TEXTCOLOR", (0, 0), (0, -1), colors.white),  # Texto blanco en columna "N°."

                ("GRID", (0, 0), (-1, -1), 1, colors.black)
            ]))


            elementos.append(tabla)
            doc.build(elementos)

            messagebox.showinfo("Éxito", "PDF generado correctamente en la carpeta del proyecto.")
        except Exception as e:
            mostrar_error_conexion_si_aplica(e, "generar el PDF")

    # Botón para generar PDF
    btn_pdf = tk.Button(frame, text="Generar y Guardar PDF", command=generar_pdf,
                        font=("Sans-serif", 18, "bold"), bg="#607D8B", fg="white", height=2)
    btn_pdf.pack(pady=10)

    return frame